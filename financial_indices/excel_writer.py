from abc import (ABCMeta,
                 abstractmethod)
import datetime
import logging
import os
from types import MappingProxyType
from typing import (Dict,
                    Collection,
                    Iterable,
                    Optional,
                    Tuple,
                    Union,
                    )

import openpyxl as xlsx

from bcb_api import (DAY_RECORD,
                     IndicesRecord,
                     RECORDS)


logger = logging.getLogger(__name__)


class WorksheetWriter(metaclass=ABCMeta):
    """ Class responsible for writing information in a Worksheet object."""

    def __init__(self, worksheet: 'openpyxl.worksheet.worksheet.Worksheet',
                 records: RECORDS):
        """
        :param worksheet:
        :param records:
        """

        self._worksheet = worksheet
        self._indices_records = records
        self._headers = self._get_headers()

        self._write_records()

    def _get_headers(self) -> Tuple[Union[str, float]]:
        """ Return the name of the headers. Headers may be text or numbers.
        """

        return ('date', 'value',)

    @abstractmethod
    def _format_record(self, record: DAY_RECORD) -> Collection:
        """ Return a collection of values corresponding to a row of data."""
        pass

    def _write_headers(self) -> None:
        """ Write the self._headers values starting at row 1, column 1."""

        for column, header in enumerate(self._headers, 1):
            self._worksheet.cell(1, column).value = header

    def _get_first_row(self, first_date: datetime.date) -> int:
        """ Return the first row to start writing the self._indices_records,
        based on the first_date value provided.

        :param first_date: The first date from self._indices_records.
        :return: Integer of the row to start writing.
        """

        column = 1
        for row in range(self._worksheet.max_row, 1, -1):
            date_on_cell = self._worksheet.cell(row, column).value.date()

            if date_on_cell == first_date:
                return row
            elif date_on_cell < first_date:
                return row + 1
        else:
            return 1

    def _erase_extra_records(self, row: int) -> None:
        """ Removes all records from row to the max row, and column 1 to the
        max column.

        :param row: Integer of the first row to erase values.
        :return: None.
        """

        for row in range(row, self._worksheet.max_row + 1):
            for column in range(1, self._worksheet.max_column + 1):
                self._worksheet.cell(row, column).value = None

    def _write_records(self) -> None:
        """ Write all dates and values from self._indices_records in
        self._worksheet.

        :return: None.
        """

        try:
            first_date = self._indices_records[0].date
        except KeyError:
            first_row = 1
        else:
            first_row = self._get_first_row(first_date)

        if first_row == 1:
            self._write_headers()
            first_row = 2

        for row, record in enumerate(self._indices_records, first_row):
            formatted_record = self._format_record(record)
            for column, column_data in enumerate(formatted_record, 1):
                self._worksheet.cell(row, column).value = column_data

        try:
            self._erase_extra_records(row + 1)
        except NameError:
            self._erase_extra_records(first_row)


class SelicWriter(WorksheetWriter):

    def _format_record(self, record: DAY_RECORD
                       ) -> Iterable[Union[float, datetime.date]]:
        """ Format a record to be appropriate to the worksheet selic.

        :param record: IndicesRecord.
        :return: Iterable of the values of record.
        """

        return record.date, record.value


class CdiWriter(WorksheetWriter):

    @staticmethod
    def _decimal_precision(value: float) -> int:
        """ Return an integer signifying how many decimal point are necessary
        to properly represent value.

        :param value: Any number.
        :return: Necessary decimal points to represent value.
        """

        precision = 0
        while not float.is_integer(value):
            value *= 10
            precision += 1

        return precision

    def _generate_percentage_range(self, start: float, end: float, step: float
                                   ) -> Tuple[float]:
        """ Return all values between start and end (both included), that can
        can be achieved with increments of step.

        Pre-Condition: (start + (step * number_steps)) == end

        :param start: Initial value.
        :param end: Last value.
        :param step: Increment step.
        :return: All values between start and end (included).
        """

        number_steps = (end - start) / step

        # If number_steps isn't an integer, than it's impossible for the
        # summation of start and steps to ever reach the end.
        if not float.is_integer(number_steps):
            raise ValueError(f'Wrong values for start: {start} end:{end} '
                             f'and step:{step}')
        else:
            number_steps = int(number_steps)

        decimal_points = self._decimal_precision(step)
        cdi_range = [start]
        for _ in range(number_steps):
            start = round(start + step, decimal_points)
            cdi_range.append(start)

        assert cdi_range[-1] == end

        return tuple(cdi_range)

    def _get_headers(self) -> Tuple[Union[str, float]]:
        """ Generate the headers for a CDI worksheet, with percentages
        ranging from 0.700 to 2.000, with increments of 0.001 (70% to 200%).

        :return: Tuple of values.
        """

        return super()._get_headers() + self._generate_percentage_range(
            0.7, 2.0, 0.001
        )

    def _format_record(self, record: DAY_RECORD
                       ) -> Iterable[Union[float, datetime.date]]:
        """ Format a record to be appropriate to the worksheet cdi.

        :param record: IndicesRecord.
        :return: Iterable of the values of record.
        """

        elements = [record.date, record.value]
        for percentage in self._headers[len(super()._get_headers()):]:
            elements.append(1 + round(record.value * percentage / 100, 8))

        return iter(elements)


class IpcaWriter(WorksheetWriter):

    def _get_headers(self) -> Tuple[Union[str, float]]:
        """ Return the name of the headers. Headers may be string or numbers.
        """

        return ('ano', 'mes', 'valor')

    def _format_record(self, record: DAY_RECORD
                       ) -> Iterable[Union[float, datetime.date]]:
        """ Format a record to be appropriate to the worksheet ipca.

        :param record: IndicesRecord.
        :return: Iterable of the values of record.
        """

        return record.date.year, record.date.month, record.value

    def _get_first_row(self, first_date: datetime.date) -> int:
        """ Return the first row to start writing the self._indices_records,
        based on the first_date value provided.

        :param first_date: The first date from self._indices_records.
        :return: Integer of the row to start writing.
        """

        for row in range(self._worksheet.max_row, 1, -1):
            year = self._worksheet.cell(row, 1).value
            month = self._worksheet.cell(row, 2).value
            date_on_cell = datetime.date(year, month, 1)

            if date_on_cell == first_date:
                return row
            elif date_on_cell < first_date:
                return row + 1
        else:
            return 1


class TrWriter(WorksheetWriter):

    def _get_headers(self) -> Tuple[Union[str, float]]:
        """ Return the name of the headers. Headers may be string or numbers.
        """

        return ('data inicial', 'data final', 'valor')

    def _format_record(self, record: DAY_RECORD) -> Collection:
        """ Format a record to be appropriate to the worksheet tr.

        :param record: IndicesRecord.
        :return: Iterable of the values of record.
        """

        return record.date, record.end_date, record.value


class MetadataWriter:
    """ Class to write and update values in the worksheet responsible for storing
    metadata information about each indices.
    """

    def __init__(self, worksheet: 'openpyxl.worksheet.worksheet.Worksheet'):
        """ Constructor of MetadataWriter."""

        self._worksheet = worksheet

        self._write_headers()
        self.indices_dates = self._get_indices_last_date()

    def _write_headers(self) -> None:
        """ Write the the header values starting at row 1, column 1."""

        for column, header in enumerate(('indices', 'last date'), 1):
            self._worksheet.cell(1, column).value = header

    def _get_indices_last_date(self) -> Dict[int, Union[datetime.date, None]]:
        """ Return a dictionary with all existing indices and last dates stored
        in self._worksheet.
        """
        indices_date = {}
        for row in range(self._worksheet.max_row, 1, -1):
            try:
                cod = int(self._worksheet.cell(row, 1).value)
            except TypeError:
                continue
            try:
                date = self._worksheet.cell(row, 2).value.date()
            except AttributeError:
                date = None
            indices_date[cod] = date

        return indices_date

    def write_indices_last_date(self) -> None:
        """ Writes self.indices_dates values on self._worksheet."""

        row = 2
        for indices, date in sorted(self.indices_dates.items()):
            self._worksheet.cell(row, 1).value = indices
            self._worksheet.cell(row, 2).value = date
            row += 1


class IndicesWorkbook:
    """ Class to represent an excel Workbook."""

    _worksheet_properties = MappingProxyType(
        {
            -1: {
                'name': 'metadata',
                'color': '000000',
                'writer': MetadataWriter,
                'state': 'veryHidden',
            },
            11: {
                'name': 'selic',
                'color': '0000FF',  # blue
                'writer': SelicWriter,
                'state': 'visible',
            },
            12: {
                'name': 'cdi',
                'color': '00FF00',  # green
                'writer': CdiWriter,
                'state': 'visible',
            },
            433: {
                'name': 'ipca',
                'color': 'FFA500',  # orange
                'writer': IpcaWriter,
                'state': 'visible',
            },
            226: {
                'name': 'tr',
                'color': 'FF0000',  # red
                'writer': TrWriter,
                'state': 'visible',
            },
        },
    )

    def __init__(self, path_to_file: Optional[str] = None,
                 filename: str = 'financial_indices.xlsx') -> None:
        """ Constructor of a workbook.
        If path_to_file is None, than it is set to the current working directory.
        If filename exists in path_to_file, it is loaded, otherwise a new file
        is created.

        :param path_to_file: String of a valid path, where the filename exists.
        :param filename: Name of the file that either is being load or created.
        """

        if path_to_file is None:
            path_to_file = os.path.abspath(os.getcwd())

        self._workbook_path = os.path.join(path_to_file, filename)

        try:
            self._workbook = xlsx.load_workbook(self._workbook_path)
        except FileNotFoundError:
            self._workbook = xlsx.Workbook()
            self._delete_all_sheets()

        worksheet_metadata = self._create_sheet(-1)
        metadata_writer = self.__class__._worksheet_properties[-1]['writer']
        self._metadata_writer = metadata_writer(worksheet_metadata)

    def __len__(self):
        """ Return the number of worksheets inside self._workbook."""

        return len(self._workbook.sheetnames)

    def __repr__(self):
        return f'{self.__class__.__name__}({self._workbook_path})'

    def _delete_all_sheets(self) -> None:
        """ Delete all existing worksheets from self._workbook."""

        for sheet in self._workbook.sheetnames:
            del self._workbook[sheet]

    def _create_sheet(self, indices_code: int
                      ) -> 'openpyxl.worksheet.worksheet.Worksheet':
        """ Create and return a worksheet in self._workbook based on the
        indices_code value. If that indices_code worksheet already exists, it
        is simply returned.

        :param indices_code: Integer representing a financial indices.
        :return: Worksheet object.
        """

        name = self.__class__._worksheet_properties[indices_code]['name']
        try:
            return self._workbook[name]
        except KeyError:
            color = self.__class__._worksheet_properties[indices_code]['color']
            state = self.__class__._worksheet_properties[indices_code]['state']

            ws = self._workbook.create_sheet(name)
            ws.title = name
            ws.sheet_properties.tabColor = color
            ws.sheet_state = state

            return ws

    def get_last_indices_date(self, indices_code: int) -> Optional[datetime.date]:
        """ Return the date of indices_code on the self._metadata_writer.
        If indices_code value is not present in self._metadata_writer, None
        is returned.

        :param indices_code: Integer representing a financial indices.
        :return: Date or None.
        """

        try:
            return self._metadata_writer.indices_dates[indices_code]
        except KeyError:
            return None

    def write_records(self, indices_code: int,
                      records: RECORDS,
                      last_non_extended_date: Optional[datetime.date] = None) -> None:
        """ Create or load a worksheet from self._workbook, corresponding to
        the indices_code provided, and pass both the worksheet and records
        to the correct WorksheetWriter (ex: CdiWriter).

        :param indices_code: Integer representing a financial indices.
        :param records: Records of a financial indices.
        :param last_non_extended_date: The last date on records, that is a real
            record, and not an extended one.
        :return: None.
        """

        ws = self._create_sheet(indices_code)

        writer = self.__class__._worksheet_properties[indices_code]['writer']

        writer(ws, records)
        self._metadata_writer.indices_dates[indices_code] = last_non_extended_date

    def save(self) -> None:
        """ Save self._workbook at self._workbook_path."""

        self._metadata_writer.write_indices_last_date()

        self._workbook.save(self._workbook_path)
